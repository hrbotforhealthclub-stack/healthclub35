# 1. Загрузка переменных окружения должна быть самым первым действием
from dotenv import load_dotenv

load_dotenv()
from collections import defaultdict

# 2. Стандартные и сторонние библиотеки
import os
import random
import asyncio
import threading
import html
# import shutil  <-- Больше не нужен
import io
from datetime import datetime, date, time, timedelta
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from werkzeug.utils import secure_filename
from aiogram import Bot
from aiogram.client.bot import DefaultBotProperties
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# 3. Импорты из вашего проекта
from models import (
    Base, engine, get_session, Employee, Event, Idea, QuizQuestion,
    RoleOnboarding, Topic, RegCode, ArchivedEmployee, Attendance,
    ArchivedAttendance, ArchivedIdea, Role,
    BotText, OnboardingQuestion, OnboardingStep, EmployeeCustomData, RoleGuide, GroupChat,
    ConfigSetting, CircleVideo,
)

# --- НАСТРОЙКА ПРИЛОЖЕНИЯ ---
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "a_very_secret_key_for_flask")

# UPLOAD_FOLDER_* и os.makedirs* больше не нужны, т.к. всё в БД
# UPLOAD_FOLDER_ONBOARDING = 'uploads/onboarding'
# UPLOAD_FOLDER_TOPICS = 'uploads/topics'
# app.config['UPLOAD_FOLDER_ONBOARDING'] = UPLOAD_FOLDER_ONBOARDING
# app.config['UPLOAD_FOLDER_TOPICS'] = UPLOAD_FOLDER_TOPICS
# UPLOAD_FOLDER_CIRCLES = 'uploads/circles'
# app.config['UPLOAD_FOLDER_CIRCLES'] = UPLOAD_FOLDER_CIRCLES
# os.makedirs(UPLOAD_FOLDER_CIRCLES, exist_ok=True)
# os.makedirs(UPLOAD_FOLDER_ONBOARDING, exist_ok=True)
# os.makedirs(UPLOAD_FOLDER_TOPICS, exist_ok=True)


# Эта команда создаст таблицы, если их нет. Для изменений используйте миграции Alembic.
Base.metadata.create_all(engine)

ADMIN_USERNAME = os.getenv("ADMIN_USERNAME")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")
if not ADMIN_USERNAME or not ADMIN_PASSWORD:
    raise RuntimeError("ADMIN_USERNAME/ADMIN_PASSWORD не заданы в .env — задайте перед запуском.")

ONBOARDING_DATA_KEYS = {
    'name': 'Имя (обновляет основной профиль)',
    'birthday': 'Дата рождения (обновляет профиль, формат ДД.ММ.ГГГГ)',
    'contact_info': 'Контактная информация (обновляет профиль)',
    'hobby': 'Хобби (дополнительное поле)',
    'favorite_quote': 'Любимая цитата (дополнительное поле)',
    'tshirt_size': 'Размер футболки (дополнительное поле)'
}

# --- ХЕЛПЕРЫ ДЛЯ КОНФИГУРАЦИИ В БД ---

# простой кэш в памяти процесса
CONFIG_CACHE: dict[str, str] = {}


# make_circle_filename больше не нужен
# def make_circle_filename(original_filename: str) -> str: ...


def get_config_value(key: str, default: str = "") -> str:
    """Получает значение настройки из БД. Создает с default, если не найдено."""
    with get_session() as db:
        setting = db.get(ConfigSetting, key)
        if setting and setting.value is not None:
            return setting.value

        if setting:
            setting.value = default
        else:
            setting = ConfigSetting(key=key, value=default)
            db.add(setting)
        db.commit()
        return default


def _collect_upcoming_birthdays(employees, days_ahead: int = 7):
    """Вернёт список сотрудников, у кого ДР в ближайшие days_ahead дней."""
    today = date.today()
    result = []
    for emp in employees:
        if not getattr(emp, "birthday", None):
            continue

        bd = emp.birthday  # date из БД

        # переносим на этот год
        try:
            this_year_bd = bd.replace(year=today.year)
        except ValueError:
            # 29 февраля
            this_year_bd = bd.replace(year=today.year, day=28)

        # если уже прошло — в следующий год
        if this_year_bd < today:
            try:
                this_year_bd = this_year_bd.replace(year=today.year + 1)
            except ValueError:
                this_year_bd = this_year_bd.replace(year=today.year + 1, day=28)

        delta_days = (this_year_bd - today).days
        if 0 <= delta_days <= days_ahead:
            result.append({
                "id": emp.id,
                "name": emp.name,
                "role": getattr(emp, "role", None),
                "date": this_year_bd,
                "in_days": delta_days,
            })

    # чтобы сначала были "сегодня/завтра"
    result.sort(key=lambda x: x["in_days"])
    return result


def save_employee_custom_field(employee_id: int, data_key: str, data_value: str):
    """Создаёт или обновляет кастомное поле сотрудника (онбординг)."""
    if not data_key:
        return
    with get_session() as db:
        obj = (
            db.query(EmployeeCustomData)
            .filter_by(employee_id=employee_id, data_key=data_key)
            .first()
        )
        if obj:
            obj.data_value = data_value
        else:
            obj = EmployeeCustomData(
                employee_id=employee_id,
                data_key=data_key,
                data_value=data_value
            )
            db.add(obj)
        db.commit()


def get_config_cached(key: str, default: str = "") -> str:
    """Быстрый вариант: сначала смотрим в память, потом в БД."""
    if key in CONFIG_CACHE:
        return CONFIG_CACHE[key]
    val = get_config_value(key, default)
    CONFIG_CACHE[key] = val
    return val


def set_config_value(key: str, value: str):
    """Устанавливает значение настройки в БД."""
    with get_session() as db:
        setting = db.get(ConfigSetting, key)
        if setting:
            setting.value = value
        else:
            setting = ConfigSetting(key=key, value=value)
            db.add(setting)
        db.commit()
    # обновляем кэш, чтобы notify_common_chat видел актуальные чаты
    CONFIG_CACHE[key] = value


# --- ОБЩИЕ ХЕЛПЕРЫ И ФИЛЬТРЫ ---
@app.template_filter('fmt_dt')
def fmt_dt(value, fmt='%Y-%m-%d %H:%M'):
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        try:
            return value.strftime(fmt)
        except Exception:
            pass
    s = str(value)
    patterns = [
        '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M',
        '%Y-%m-%d %H:%M', '%Y-%m-%d', '%d.%m.%Y %H:%M', '%d.%m.%Y',
    ]
    for p in patterns:
        try:
            dt = datetime.strptime(s, p)
            return dt.strftime(fmt)
        except Exception:
            continue
    return s


def _run_async_bg(coro):
    """Запуск корутины в отдельном event loop в фоне, всегда стабильно."""

    def runner():
        new_loop = asyncio.new_event_loop()
        asyncio.set_event_loop(new_loop)
        new_loop.run_until_complete(coro)

    threading.Thread(target=runner, daemon=True).start()


@app.template_filter('fmt_date')
def fmt_date(value, fmt="%d.%m.%Y"):
    if not value: return ""
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return value
    return value.strftime(fmt)


@app.template_filter('fmt_time')
def fmt_time(value, fmt='%H:%M:%S'):
    if value is None: return ''
    return value.strftime(fmt)


# --- АУТЕНТИФИКАЦИЯ И ДОСТУП ---
def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)

    return wrapped


@app.before_request
def require_login_for_all():
    open_paths = {"/login", "/logout", "/landing"}
    # Добавляем маршруты для файлов, если хотим, чтобы они были видны без логина
    # (но лучше оставить @login_required на самих маршрутах)
    if request.path.startswith("/static"):
        return
    if request.path in open_paths:
        return
    if not session.get("is_admin"):
        return redirect(url_for("login", next=request.path))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["is_admin"] = True
            flash("Вы вошли в админ-панель.", "success")
            next_url = request.args.get("next") or url_for("index")
            return redirect(next_url)
        else:
            error = "Неверный логин или пароль."
    return render_template("login.html", error=error)


@app.route("/logout", methods=["POST", "GET"])
def logout():
    session.pop("is_admin", None)
    flash("Вы вышли.", "info")
    return redirect(url_for("login"))


# --- ЛОГИКА TELEGRAM ---
def get_text(key: str, default: str = "Текст не найден") -> str:
    """Получает текст для бота из БД по ключу."""
    with get_session() as db:
        text_obj = db.get(BotText, key)
        return text_obj.text if text_obj else default


def _chat_candidates(raw: str | int) -> list:
    raw = str(raw if raw is not None else "").strip()
    if not raw: return []
    if raw.startswith("@"): return [raw]
    if raw.startswith("-100"): return [raw]
    if raw.startswith("-") and raw[1:].isdigit(): return [raw, f"-100{raw[1:]}"]
    return [raw]


async def _send_tg_message(text: str, chat_id: str):
    token = get_config_cached("BOT_TOKEN")
    if not token:
        print("[tg] BOT_TOKEN missing in DB")
        return False, "BOT_TOKEN missing"
    if not chat_id:
        print(f"[tg] chat_id is missing for text: {text[:50]}...")
        return False, "chat_id is missing"

    target = int(chat_id) if str(chat_id).strip().lstrip('-').isdigit() else chat_id
    bot = Bot(token=token, default=DefaultBotProperties(parse_mode="HTML"))
    try:
        msg = await bot.send_message(chat_id=target, text=text)
        return True, f"sent:{msg.message_id}"
    except Exception as e:
        print(f"[tg] send error to {chat_id}: {e}")
        return False, str(e)
    finally:
        await bot.session.close()


def notify_common_chat(text: str):
    async def _send_to_all_active_chats():
        active_chats_str = get_config_cached("ACTIVE_CHAT_IDS", "")

        if not active_chats_str:
            print("[tg] ACTIVE_CHAT_IDS is not set in DB")
            return
        chat_ids = [cid.strip() for cid in active_chats_str.split(',') if cid.strip()]
        tasks = [_send_tg_message(text, chat_id) for chat_id in chat_ids]
        if tasks: await asyncio.gather(*tasks)

    _run_async_bg(_send_to_all_active_chats())


async def _list_verified_admin_groups_async(rows):
    token = get_config_value("BOT_TOKEN")
    if not token: return []
    bot = Bot(token=token, default=DefaultBotProperties(parse_mode="HTML"))
    out = []
    try:
        me = await bot.get_me()
        for row in rows:
            for cid in _chat_candidates(row.chat_id):
                try:
                    chat = await bot.get_chat(cid)
                    mem = await bot.get_chat_member(chat.id, me.id)
                    status = str(getattr(mem, "status", ""))
                    if "administrator" in status or "creator" in status:
                        norm_id = str(chat.id)
                        name = getattr(chat, "title", None) or getattr(chat, "username", None) or row.name
                        out.append({"chat_id": norm_id, "name": name, "_db_id": getattr(row, "id", None)})
                        break
                except Exception:
                    continue
        return out
    finally:
        await bot.session.close()


def list_admin_groups_from_db_only():
    """Быстрый вариант: читаем только то, что уже есть в БД, без запросов к Telegram."""
    with get_session() as db:
        rows = db.query(GroupChat).order_by(GroupChat.name).all()
        return [
            {
                "chat_id": r.chat_id,
                "name": (r.name or getattr(r, "title", None) or r.username or r.chat_id)
            }
            for r in rows
        ]


@app.route('/')
def index():
    # УЛУЧШЕНИЕ: Проверяем токен при загрузке главной страницы
    bot_token = get_config_cached("BOT_TOKEN")
    if not bot_token:
        flash(
            "Внимание: Токен бота не задан. Пожалуйста, укажите его в разделе 'Настройки', чтобы все функции заработали.",
            "warning"
        )

    with get_session() as db:
        employees = db.query(Employee).filter_by(is_active=True).order_by(Employee.name).all()
        archived_employees = db.query(ArchivedEmployee).order_by(ArchivedEmployee.dismissal_date.desc()).all()
        events = db.query(Event).order_by(Event.event_date.desc()).all()
        ideas = (
            db.query(Idea, Employee.name)
            .outerjoin(Employee, Idea.employee_id == Employee.id)
            .order_by(Idea.submission_date.desc())
            .all()
        )
        topics = db.query(Topic).order_by(Topic.title).all()
        roles = db.query(Role).order_by(Role.name).all()
        bot_texts = db.query(BotText).order_by(BotText.id).all()
        attendance_records = (
            db.query(Attendance, Employee.name)
            .join(Employee, Attendance.employee_id == Employee.id)
            .order_by(Attendance.date.desc(), Attendance.arrival_time.desc())
            .all()
        )

        # ВАЖНО: тащим всё разом, а не по роли
        all_questions = (
            db.query(OnboardingQuestion)
            .order_by(OnboardingQuestion.role, OnboardingQuestion.order_index)
            .all()
        )
        all_steps = (
            db.query(OnboardingStep)
            .order_by(OnboardingStep.role, OnboardingStep.order_index)
            .all()
        )
        all_guides = (
            db.query(RoleGuide)
            .order_by(RoleGuide.role, RoleGuide.order_index)
            .all()
        )
        all_custom_data = (
            db.query(EmployeeCustomData)
            .filter(EmployeeCustomData.employee_id.isnot(None))
            .all()
        )

        all_onboarding_infos = db.query(RoleOnboarding).all()
        all_quizzes = (
            db.query(QuizQuestion)
            .order_by(QuizQuestion.role, QuizQuestion.order_index)
            .all()
        )

        # <<< НОВОЕ: список ДР на 7 дней вперёд >>>
        upcoming_birthdays = _collect_upcoming_birthdays(employees, days_ahead=7)

    # группируем в питоне
    questions_by_role = defaultdict(list)
    for q in all_questions:
        questions_by_role[q.role].append(q)

    steps_by_role = defaultdict(list)
    for s in all_steps:
        steps_by_role[s.role].append(s)

    guides_by_role = defaultdict(list)
    for g in all_guides:
        guides_by_role[g.role].append(g)

    onboarding_info_by_role = {r.role: r for r in all_onboarding_infos}

    quizzes_by_role = defaultdict(list)
    for q in all_quizzes:
        quizzes_by_role[q.role].append(q)

    custom_data_by_employee = defaultdict(dict)
    for cd in all_custom_data:
        if not cd:
            continue
        if not cd.employee_id:
            continue
        if not cd.data_key:
            continue
        custom_data_by_employee[cd.employee_id][cd.data_key] = cd.data_value or ""

    onboarding_constructor_data = {}
    onboarding_data = {}
    role_guides_data = {}

    for role in roles:
        rname = role.name
        onboarding_constructor_data[rname] = {
            "questions": questions_by_role.get(rname, []),
            "steps": steps_by_role.get(rname, []),
        }
        onboarding_data[rname] = {
            "info": onboarding_info_by_role.get(rname),
            "quizzes": quizzes_by_role.get(rname, []),
        }
        role_guides_data[rname] = guides_by_role.get(rname, [])

    # конфиг — из кэша, чтобы не бить БД 4 раза
    config = {
        "BOT_TOKEN": bot_token,
        "ACTIVE_CHAT_IDS": [
            c.strip()
            for c in (get_config_cached("ACTIVE_CHAT_IDS", "") or "").split(',')
            if c.strip()
        ],
        "OFFICE_LAT": get_config_cached("OFFICE_LAT", ""),
        "OFFICE_LON": get_config_cached("OFFICE_LON", ""),
        "OFFICE_RADIUS_METERS": get_config_cached("OFFICE_RADIUS_METERS", "")
    }

    # чаты — быстрый вариант
    admin_groups = list_admin_groups_from_db_only()

    return render_template(
        'index.html',
        employees=employees,
        archived_employees=archived_employees,
        events=events,
        ideas=ideas,
        topics=topics,
        onboarding_data=onboarding_data,
        roles=roles,
        config=config,
        bot_texts=bot_texts,
        onboarding_constructor_data=onboarding_constructor_data,
        onboarding_data_keys=ONBOARDING_DATA_KEYS,
        attendance_records=attendance_records,
        role_guides_data=role_guides_data,
        admin_groups=admin_groups,
        custom_data_by_employee=custom_data_by_employee,
        # <<< НОВОЕ: отдаём в шаблон >>>
        upcoming_birthdays=upcoming_birthdays,
    )


# --- AJAX-МАРШРУТЫ (CRUD) ---

@app.route('/texts/update/<string:text_id>', methods=['POST'])
def update_text(text_id):
    with get_session() as db:
        text_obj = db.get(BotText, text_id)
        if text_obj:
            text_obj.text = request.form.get('text', '')
            db.commit()
            return jsonify({"success": True, "message": f"Текст '{text_id}' обновлен.", "category": "success"})
        return jsonify({"success": False, "message": "Текст не найден.", "category": "danger"}), 404


@app.route('/onboarding/question/add/<path:role>', methods=['POST'])
def add_onboarding_question(role):
    with get_session() as db:
        max_idx = db.query(func.max(OnboardingQuestion.order_index)).filter_by(role=role).scalar()
        next_idx = (max_idx + 1) if max_idx is not None else 0
        new_q = OnboardingQuestion(
            role=role,
            question_text=request.form['question_text'],
            data_key=request.form['data_key'],
            is_required=('is_required' in request.form),
            order_index=next_idx
        )
        db.add(new_q)
        db.commit()
        return jsonify({
            "success": True, "message": "Новый вопрос для онбординга добавлен.", "category": "success",
            "item": {
                "id": new_q.id, "question_text": new_q.question_text, "data_key": new_q.data_key,
                "is_required": new_q.is_required, "delete_url": url_for('delete_onboarding_question', q_id=new_q.id)
            }
        })


@app.route('/onboarding/question/delete/<int:q_id>', methods=['POST'])
def delete_onboarding_question(q_id):
    with get_session() as db:
        q = db.get(OnboardingQuestion, q_id)
        if q:
            db.delete(q)
            db.commit()
            return jsonify({"success": True, "message": "Вопрос онбординга удален.", "category": "warning"})
    return jsonify({"success": False, "message": "Вопрос не найден.", "category": "danger"}), 404


@app.route('/onboarding/question/reorder', methods=['POST'])
def reorder_onboarding_question():
    ordered_ids = request.get_json(silent=True).get('ordered_ids', [])
    if not ordered_ids:
        return jsonify(success=False, message="Нет данных для сортировки"), 400
    with get_session() as session:
        questions_map = {
            str(q.id): q for q in session.query(OnboardingQuestion).filter(
                OnboardingQuestion.id.in_([int(i) for i in ordered_ids if i.isdigit()])
            ).all()
        }
        for index, qid in enumerate(ordered_ids):
            if qid in questions_map:
                questions_map[qid].order_index = index
        session.commit()
    return jsonify(success=True)


@app.route('/onboarding/step/add/<path:role>', methods=['POST'])
def add_onboarding_step(role):
    with get_session() as db:

        # v-- ЭТИ СТРОКИ НУЖНО ДОБАВИТЬ --v
        max_idx = db.query(func.max(OnboardingStep.order_index)).filter_by(role=role).scalar()
        next_idx = (max_idx + 1) if max_idx is not None else 0
        # ^-- КОНЕЦ НОВОГО КОДА --^

        new_step = OnboardingStep(
            role=role,
            message_text=request.form.get('message_text'),
            file_type=request.form.get('file_type', 'document'),
            order_index=next_idx  # <-- И ОБЯЗАТЕЛЬНО ДОБАВИТЬ ЭТО ПОЛЕ
        )

        file_url = None
        if 'file' in request.files and request.files['file'].filename != '':
            file = request.files['file']
            # Читаем файл в память и сохраняем в БД
            new_step.file_data = file.read()
            new_step.file_mime = file.mimetype
            new_step.file_name = secure_filename(file.filename)

        db.add(new_step)
        db.commit()  # commit, чтобы получить new_step.id

        if new_step.file_data:
            file_url = url_for('serve_onboarding_step_file', step_id=new_step.id)

        return jsonify({
            "success": True, "message": "Новый шаг знакомства добавлен.", "category": "success",
            "item": {
                "id": new_step.id, "message_text": new_step.message_text, "file_type": new_step.file_type,
                # "file_path": new_step.file_path, <-- Убрали
                "file_url": file_url,  # <-- Добавили
                "file_name": new_step.file_name,
                "delete_url": url_for('delete_onboarding_step', step_id=new_step.id)
            }
        })


@app.route('/onboarding/step/delete/<int:step_id>', methods=['POST'])
def delete_onboarding_step(step_id):
    with get_session() as db:
        step = db.get(OnboardingStep, step_id)
        if step:
            # if step.file_path and os.path.exists(step.file_path): <-- Больше не нужно
            #     os.remove(step.file_path) <-- Больше не нужно
            db.delete(step)
            db.commit()
            return jsonify({"success": True, "message": "Шаг знакомства удален.", "category": "warning"})
    return jsonify({"success": False, "message": "Шаг не найден.", "category": "danger"}), 404


@app.post("/api/onboarding/save_custom_data")
def api_onboarding_save_custom_data():
    if not session.get("is_admin") and request.headers.get("X-Internal-Token") != os.getenv("INTERNAL_BOT_TOKEN", ""):
        # если хочешь, можешь вообще убрать эту проверку
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    payload = request.get_json(silent=True) or {}
    employee_id = payload.get("employee_id")
    data = payload.get("data") or {}

    if not employee_id or not isinstance(data, dict):
        return jsonify({"ok": False, "error": "bad payload"}), 400

    # сохраняем все поля, которые пришли
    for key, value in data.items():
        if value is None:
            value = ""
        save_employee_custom_field(employee_id, key, str(value))

        # спец-логика: некоторые ключи должны обновлять основной профиль
        if key == "name" and value:
            with get_session() as db:
                emp = db.get(Employee, employee_id)
                if emp:
                    emp.name = value
                    db.commit()
        if key == "birthday" and value:
            # принимаем и 2025-10-31 и 31.10.2025
            parsed = None
            for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
                try:
                    parsed = datetime.strptime(value, fmt).date()
                    break
                except Exception:
                    continue
            if parsed:
                with get_session() as db:
                    emp = db.get(Employee, employee_id)
                    if emp:
                        emp.birthday = parsed
                        db.commit()

    return jsonify({"ok": True})


@app.route('/onboarding/step/reorder', methods=['POST'])
def reorder_onboarding_step():
    ordered_ids = request.get_json(silent=True).get('ordered_ids', [])
    with get_session() as session:
        for index, sid in enumerate(ordered_ids):
            step = session.get(OnboardingStep, int(sid))
            if step:
                step.order_index = index
        session.commit()
    return jsonify(success=True)


@app.route('/role/add', methods=['POST'])
def add_role():
    with get_session() as db:
        role_name = request.form.get('role_name')
        if not role_name:
            return jsonify({"success": False, "message": "Имя роли не указано.", "category": "danger"}), 400
        if db.query(Role).filter_by(name=role_name).first():
            return jsonify(
                {"success": False, "message": f"Роль '{role_name}' уже существует.", "category": "danger"}), 409
        new_role = Role(name=role_name)
        db.add(new_role)
        db.commit()
        return jsonify({
            "success": True, "message": f"Роль '{role_name}' добавлена.", "category": "success",
            "role": {"id": new_role.id, "name": new_role.name}
        })


@app.route('/role/delete/<int:role_id>', methods=['POST'])
def delete_role(role_id):
    with get_session() as db:
        role = db.get(Role, role_id)
        if role:
            db.delete(role)
            db.commit()
            return jsonify({"success": True, "message": f"Роль '{role.name}' удалена.", "category": "warning"})
    return jsonify({"success": False, "message": "Роль не найдена.", "category": "danger"}), 404


@app.route('/employee/add', methods=['POST'])
def add_employee():
    with get_session() as db:
        email, role = request.form.get('email'), request.form.get('role')
        if not email or not role:
            return jsonify({"success": False, "message": "Email и роль обязательны.", "category": "danger"}), 400
        if db.query(Employee).filter_by(email=email).first():
            return jsonify(
                {"success": False, "message": f"Сотрудник с email {email} уже существует.", "category": "danger"}), 409

        new_emp = Employee(email=email, name=email, role=role, is_active=True)
        db.add(new_emp)
        db.commit()

        return jsonify({
            "success": True,
            "message": f"Сотрудник с email {email} добавлен.",
            "category": "success",
            "action": "reload"  # <-- ИЗМЕНЕНИЕ: Говорим фронтенду перезагрузить страницу
        })


@app.route('/employee/reset_progress/<int:emp_id>', methods=['POST'])
def reset_progress(emp_id):
    with get_session() as db:
        emp = db.get(Employee, emp_id)
        if not emp:
            return jsonify({"success": False, "message": "Сотрудник не найден.", "category": "danger"}), 404
        emp.onboarding_completed = False
        emp.training_passed = False
        db.query(EmployeeCustomData).filter_by(employee_id=emp_id).delete(synchronize_session=False)
        db.commit()
        if emp.telegram_id:
            _run_async_bg(
                _send_tg_message(get_text('progress_reset_notification', 'Ваш прогресс сброшен.'), emp.telegram_id))
        return jsonify({"success": True, "message": f"Прогресс для {emp.name} сброшен.", "category": "warning"})


@app.route('/broadcast/send', methods=['POST'])
def send_broadcast():
    message_text, target_role = request.form.get('message_text'), request.form.get('target_role')
    if not message_text:
        return jsonify(
            {"success": False, "message": "Текст сообщения не может быть пустым.", "category": "danger"}), 400
    with get_session() as db:
        query = db.query(Employee.telegram_id).filter(Employee.is_active == True, Employee.telegram_id != None)
        if target_role != 'all': query = query.filter(Employee.role == target_role)
        target_users_ids = [row[0] for row in query.all()]
    if not target_users_ids:
        return jsonify({"success": False, "message": "Не найдено сотрудников для рассылки.", "category": "warning"})

    async def _send_to_all():
        token = get_config_value("BOT_TOKEN")
        if not token: return
        bot = Bot(token=token)
        try:
            for user_id in target_users_ids:
                try:
                    await bot.send_message(chat_id=user_id, text=message_text)
                except Exception:
                    pass
                await asyncio.sleep(0.1)
        finally:
            await bot.session.close()

    _run_async_bg(_send_to_all())
    return jsonify({"success": True, "message": f"Рассылка для {len(target_users_ids)} пользователей запущена.",
                    "category": "info"})


@app.route('/employee/edit/<int:emp_id>', methods=['POST'])
def edit_employee(emp_id):
    with get_session() as db:
        emp = db.get(Employee, emp_id)
        if not emp:
            return jsonify({"success": False, "message": "Сотрудник не найден.", "category": "danger"}), 404

        old_role = emp.role  # что было до изменения

        emp.name = request.form.get('name', emp.name)
        emp.email = request.form.get('email', emp.email)
        emp.role = request.form.get('role', emp.role)
        birthday_str = request.form.get('birthday')
        emp.birthday = datetime.strptime(birthday_str, '%Y-%m-%d').date() if birthday_str else None
        db.commit()

    # вне сессии: если роль реально изменилась — шлём в чаты
    if old_role != emp.role:
        raw_tpl = get_text(
            'employee_role_changed_announcement',
            '⬆️ {name} был(а) повышен(а): {old} → {new}'
        )

        # делаем словарь со ВСЕМИ часто встречающимися ключами
        data = {
            "name": html.escape(emp.name or emp.email or "Сотрудник"),
            "old": html.escape(old_role or "—"),
            "new": html.escape(emp.role or "—"),
            "old_role": html.escape(old_role or "—"),
            "new_role": html.escape(emp.role or "—"),
        }

        # безопасное форматирование: если в шаблоне есть левый {xxx}, он станет пустым
        class SafeDict(dict):
            def __missing__(self, key):
                return ""

        msg = raw_tpl.format_map(SafeDict(data))
        notify_common_chat(msg)

    return jsonify({
        "success": True,
        "message": f"Данные сотрудника {emp.name} обновлены.",
        "category": "success",
        "employee": {"id": emp.id, "name": emp.name, "email": emp.email, "role": emp.role}
    })


@app.route('/employee/dismiss/<int:emp_id>', methods=['POST'])
def dismiss_employee(emp_id):
    with get_session() as db:
        emp = db.get(Employee, emp_id)
        if not emp:
            return jsonify({"success": False, "message": "Сотрудник не найден.", "category": "danger"}), 404
        name_cache, role_cache = emp.name or emp.email, emp.role or ''
        emp.is_active = False
        emp.registered = False
        emp.telegram_id = None
        db.commit()
        dismiss_text = get_text('employee_dismissed_announcement', '👋 {name} ({role}) больше не с нами.').format(
            name=html.escape(name_cache), role=html.escape(role_cache))
        notify_common_chat(dismiss_text)
        return jsonify({"success": True, "message": f"Сотрудник {name_cache} деактивирован.", "category": "warning"})


@app.route('/employee/reset_telegram/<int:emp_id>', methods=['POST'])
def reset_telegram(emp_id):
    with get_session() as db:
        emp = db.get(Employee, emp_id)
        if not emp: return jsonify({"success": False, "message": "Сотрудник не найден.", "category": "danger"}), 404
        emp.telegram_id = None
        emp.registered = False
        db.commit()
        return jsonify(
            {"success": True, "message": f"Привязка Telegram для {emp.name} сброшена.", "category": "warning"})


@app.route('/employee/generate_code/<int:emp_id>', methods=['POST'])
def generate_new_code(emp_id):
    with get_session() as db:
        emp = db.get(Employee, emp_id)
        if not emp: return jsonify({"success": False, "message": "Сотрудник не найден.", "category": "danger"}), 404
        while True:
            code = "".join(str(random.randint(0, 9)) for _ in range(8))
            if not db.query(RegCode).filter_by(code=code).first(): break
        db.add(RegCode(code=code, email=emp.email, used=False))
        db.commit()
        return jsonify({"success": True, "message": f"Новый код для {emp.name}: {code}", "category": "success"})


@app.route('/onboarding/update/<path:role>', methods=['POST'])
def update_onboarding(role):
    with get_session() as db:
        onboarding = db.query(RoleOnboarding).filter_by(role=role).first()
        if not onboarding:
            onboarding = RoleOnboarding(role=role)
            db.add(onboarding)

        onboarding.text = request.form['text']
        onboarding.file_type = request.form['file_type']

        if 'file' in request.files and request.files['file'].filename != '':
            file = request.files['file']
            # Читаем в БД
            onboarding.file_data = file.read()
            onboarding.file_mime = file.mimetype
            onboarding.file_name = secure_filename(file.filename)

        db.commit()
    return jsonify({"success": True, "message": f"Онбординг для '{role}' обновлен.", "category": "success"})


@app.route('/quiz/add/<path:role>', methods=['POST'])
def add_quiz(role):
    with get_session() as db:
        qtype, question = request.form['question_type'], request.form['question']
        options = request.form.get('options') if qtype == 'choice' else None
        answer = request.form.get('answer') if qtype == 'choice' else request.form.get('text_answer')

        # --- ИСПРАВЛЕНИЕ ---
        max_idx = db.query(func.max(QuizQuestion.order_index)).filter_by(role=role).scalar()
        next_idx = (max_idx + 1) if max_idx is not None else 0

        new_q = QuizQuestion(role=role, question=question, answer=answer, question_type=qtype, options=options,
                             order_index=next_idx)  # <-- Используем next_idx
        db.add(new_q)
        db.commit()
        return jsonify({
            "success": True, "message": "Новый вопрос добавлен.", "category": "success",
            "item": {"id": new_q.id, "question": new_q.question, "answer": new_q.answer,
                     "delete_url": url_for('delete_quiz', quiz_id=new_q.id)}
        })


@app.route('/quiz/edit/<int:quiz_id>', methods=['POST'])
def edit_quiz(quiz_id):
    with get_session() as db:
        quiz = db.get(QuizQuestion, quiz_id)
        if not quiz: return jsonify({"success": False, "message": "Вопрос не найден.", "category": "danger"}), 404
        qtype = request.form['question_type']
        quiz.question = request.form['question']
        quiz.options = request.form.get('options') if qtype == 'choice' else None
        quiz.answer = request.form.get('answer') if qtype == 'choice' else request.form.get('text_answer')
        quiz.question_type = qtype
        db.commit()
        return jsonify({"success": True, "message": "Вопрос обновлен.", "category": "success"})


@app.route('/quiz/delete/<int:quiz_id>', methods=['POST'])
def delete_quiz(quiz_id):
    with get_session() as db:
        quiz = db.get(QuizQuestion, quiz_id)
        if quiz:
            db.delete(quiz)
            db.commit()
            return jsonify({"success": True, "message": "Вопрос квиза удален.", "category": "warning"})
    return jsonify({"success": False, "message": "Вопрос не найден.", "category": "danger"}), 404


@app.route('/quiz/reorder', methods=['POST'])
def reorder_quiz():
    ordered_ids = request.get_json(silent=True).get('ordered_ids', [])
    with get_session() as session:
        for index, qid in enumerate(ordered_ids):
            quiz = session.get(QuizQuestion, int(qid))
            if quiz: quiz.order_index = index
        session.commit()
    return jsonify(success=True)


@app.route('/event/add', methods=['POST'])
def add_event():
    with get_session() as db:
        title, description = request.form['title'], request.form['description']
        event_dt = datetime.strptime(request.form['event_date'], '%Y-%m-%dT%H:%M')
        new_event = Event(title=title, description=description, event_date=event_dt)
        db.add(new_event)
        db.commit()
        notify_common_chat(
            get_text('event_created_announcement', '📅 {title}\n{when}\n{desc}').format(title=html.escape(title),
                                                                                       when=event_dt.strftime(
                                                                                           '%d.%m.%Y %H:%M'),
                                                                                       desc=html.escape(description)))
        return jsonify({"success": True, "message": "Новый ивент добавлен.", "category": "success", "action": "reload"})


@app.route('/event/edit/<int:event_id>', methods=['POST'])
def edit_event(event_id):
    with get_session() as db:
        event = db.get(Event, event_id)
        if not event: return jsonify({"success": False, "message": "Ивент не найден.", "category": "danger"}), 404
        event.title, event.description = request.form['title'], request.form['description']
        event.event_date = datetime.strptime(request.form['event_date'], '%Y-%m-%dT%H:%M')
        db.commit()
        return jsonify({"success": True, "message": "Ивент обновлен.", "category": "success", "action": "reload"})


@app.route('/event/delete/<int:event_id>', methods=['POST'])
def delete_event(event_id):
    with get_session() as db:
        event = db.get(Event, event_id)
        if event:
            db.delete(event)
            db.commit()
            return jsonify({"success": True, "message": "Ивент удален.", "category": "warning", "action": "reload"})
    return jsonify({"success": False, "message": "Ивент не найден.", "category": "danger"}), 404


@app.route('/idea/delete/<int:idea_id>', methods=['POST'])
def delete_idea(idea_id):
    with get_session() as db:
        idea = db.get(Idea, idea_id)
        if idea:
            db.delete(idea)
            db.commit()
            return jsonify({"success": True, "message": "Идея удалена.", "category": "warning", "action": "reload"})
    return jsonify({"success": False, "message": "Идея не найдена.", "category": "danger"}), 404


@app.route('/topic/add', methods=['POST'])
def add_topic():
    with get_session() as db:
        new_topic = Topic(title=request.form['title'], content=request.form['content'])

        if 'image' in request.files and request.files['image'].filename != '':
            file = request.files['image']
            # Читаем в БД
            new_topic.image_data = file.read()
            new_topic.image_mime = file.mimetype
            new_topic.image_name = secure_filename(file.filename)

        db.add(new_topic)
        db.commit()
    return jsonify({"success": True, "message": "Новая тема создана.", "category": "success", "action": "reload"})


@app.route('/topic/edit/<int:topic_id>', methods=['POST'])
def edit_topic(topic_id):
    with get_session() as db:
        topic = db.get(Topic, topic_id)
        if not topic: return jsonify({"success": False, "message": "Тема не найдена.", "category": "danger"}), 404

        topic.title, topic.content = request.form['title'], request.form['content']

        if 'image' in request.files and request.files['image'].filename != '':
            file = request.files['image']
            # if topic.image_path and os.path.exists(topic.image_path): os.remove(topic.image_path) <-- Убрали
            # Читаем в БД
            topic.image_data = file.read()
            topic.image_mime = file.mimetype
            topic.image_name = secure_filename(file.filename)

        db.commit()
    return jsonify({"success": True, "message": "Тема обновлена.", "category": "success", "action": "reload"})


@app.route('/topic/delete/<int:topic_id>', methods=['POST'])
def delete_topic(topic_id):
    with get_session() as db:
        topic = db.get(Topic, topic_id)
        if topic:
            # if topic.image_path and os.path.exists(topic.image_path): os.remove(topic.image_path) <-- Убрали
            db.delete(topic)
            db.commit()
            return jsonify({"success": True, "message": "Тема удалена.", "category": "warning", "action": "reload"})
    return jsonify({"success": False, "message": "Тема не найдена.", "category": "danger"}), 404


@app.route('/guide/add/<path:role>', methods=['POST'])
def add_guide(role):
    with get_session() as db:
        # --- ИСПРАВЛЕНИЕ ---
        max_idx = db.query(func.max(RoleGuide.order_index)).filter_by(role=role).scalar()
        next_idx = (max_idx + 1) if max_idx is not None else 0

        new_guide = RoleGuide(role=role, title=request.form['title'], content=request.form.get('content', ''),
                              order_index=next_idx)  # <-- Используем next_idx

        if 'file' in request.files and request.files['file'].filename != '':
            file = request.files['file']
            # Читаем в БД
            new_guide.file_data = file.read()
            new_guide.file_mime = file.mimetype
            new_guide.file_name = secure_filename(file.filename)

        db.add(new_guide)
        db.commit()
    return jsonify({"success": True, "message": "Новый регламент добавлен.", "category": "success", "action": "reload"})


@app.route('/guide/delete/<int:guide_id>', methods=['POST'])
def delete_guide(guide_id):
    with get_session() as db:
        guide = db.get(RoleGuide, guide_id)
        if guide:
            # if guide.file_path and os.path.exists(guide.file_path): os.remove(guide.file_path) <-- Убрали
            db.delete(guide)
            db.commit()
            return jsonify({"success": True, "message": "Регламент удален.", "category": "warning", "action": "reload"})
    return jsonify({"success": False, "message": "Регламент не найден.", "category": "danger"}), 404


@app.route('/config/update', methods=['POST'])
def update_config():
    active_chat_ids = request.form.getlist("ACTIVE_CHAT_IDS")
    active_chats_str = ",".join(active_chat_ids)
    set_config_value("ACTIVE_CHAT_IDS", active_chats_str)

    set_config_value("OFFICE_LAT", request.form.get("OFFICE_LAT", ""))
    set_config_value("OFFICE_LON", request.form.get("OFFICE_LON", ""))
    set_config_value("OFFICE_RADIUS_METERS", request.form.get("OFFICE_RADIUS_METERS", ""))

    new_bot_token = request.form.get("BOT_TOKEN")
    if new_bot_token:
        set_config_value("BOT_TOKEN", new_bot_token)

    return jsonify({"success": True, "message": "Настройки сохранены. Перезапустите бота, чтобы они применились.",
                    "category": "info"})


@app.route('/settings/copy', methods=['POST'])
def copy_settings():
    source_role = request.form.get('source_role')
    target_roles = request.form.getlist('target_roles')
    sections_to_copy = request.form.getlist('sections_to_copy')

    if not all([source_role, target_roles, sections_to_copy]):
        return jsonify({"success": False, "message": "Недостаточно данных для копирования.", "category": "danger"}), 400

    with get_session() as db:
        for target_role in target_roles:
            if 'scenarios' in sections_to_copy:
                db.query(OnboardingQuestion).filter_by(role=target_role).delete()
                db.query(OnboardingStep).filter_by(role=target_role).delete()

                source_questions = db.query(OnboardingQuestion).filter_by(role=source_role).order_by(
                    OnboardingQuestion.order_index).all()
                for q in source_questions:
                    new_q = OnboardingQuestion(role=target_role, question_text=q.question_text, data_key=q.data_key,
                                               is_required=q.is_required, order_index=q.order_index)
                    db.add(new_q)

                source_steps = db.query(OnboardingStep).filter_by(role=source_role).order_by(
                    OnboardingStep.order_index).all()
                for step in source_steps:
                    new_step = OnboardingStep(role=target_role, message_text=step.message_text,
                                              file_type=step.file_type, order_index=step.order_index)

                    # Копируем файл из БД
                    if step.file_data:
                        new_step.file_data = step.file_data
                        new_step.file_mime = step.file_mime
                        new_step.file_name = step.file_name

                    db.add(new_step)

            if 'training' in sections_to_copy:
                db.query(QuizQuestion).filter_by(role=target_role).delete()
                db.query(RoleOnboarding).filter_by(role=target_role).delete()

                source_quizzes = db.query(QuizQuestion).filter_by(role=source_role).order_by(
                    QuizQuestion.order_index).all()
                for quiz in source_quizzes:
                    new_quiz = QuizQuestion(role=target_role, question=quiz.question, answer=quiz.answer,
                                            question_type=quiz.question_type, options=quiz.options,
                                            order_index=quiz.order_index)
                    db.add(new_quiz)

                source_training = db.query(RoleOnboarding).filter_by(role=source_role).first()
                if source_training:
                    new_training = RoleOnboarding(role=target_role, text=source_training.text,
                                                  file_type=source_training.file_type)

                    # Копируем файл из БД
                    if source_training.file_data:
                        new_training.file_data = source_training.file_data
                        new_training.file_mime = source_training.file_mime
                        new_training.file_name = source_training.file_name

                    db.add(new_training)

        db.commit()

    return jsonify({
        "success": True,
        "message": f"Настройки из '{source_role}' скопированы в {len(target_roles)} ролей. Страница будет перезагружена.",
        "category": "success",
        "action": "reload"
    })


@app.route('/export/employees.xlsx', methods=['GET'])
def export_employees_xlsx():
    rows = []
    with get_session() as db:
        employees = db.query(Employee).order_by(Employee.name).all()
        emails = [e.email for e in employees]
        existing_free = {rc.email: rc.code for rc in
                         db.query(RegCode).filter(RegCode.used == False, RegCode.email.in_(emails)).all()}
        for emp in employees:
            code_for_row = ""
            if not emp.telegram_id:
                code_for_row = existing_free.get(emp.email, "")
                if not code_for_row:
                    while True:
                        code = "".join(str(random.randint(0, 9)) for _ in range(8))
                        if not db.query(RegCode).filter_by(code=code).first(): break
                    db.add(RegCode(code=code, email=emp.email, used=False))
                    code_for_row = code
            rows.append([
                emp.id, emp.name or "", emp.email, emp.role or "",
                emp.birthday.strftime("%d.%m.%Y") if emp.birthday else "",
                "Да" if emp.is_active else "Нет", "Да" if emp.registered else "Нет",
                "Да" if emp.training_passed else "Нет",
                        emp.telegram_id or "", code_for_row
            ])
        db.commit()

    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"
    headers = ["ID", "ФИО", "Email", "Роль", "Дата рождения", "Активен", "Зарегистрирован", "Прошёл тренинг",
               "Telegram ID", "Код для привязки"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
                         top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
    for cell in ws[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center",
                                                                                                    vertical="center"); cell.border = thin_border
    for r in rows: ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row: cell.border = thin_border
    widths = [6, 28, 28, 18, 14, 10, 16, 16, 14, 18]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.conditional_formatting.add(f"A2:{get_column_letter(len(headers))}{ws.max_row}",
                                  FormulaRule(formula=[f'LEN($I2)=0'], stopIfTrue=False,
                                              fill=PatternFill("solid", fgColor="FEE2E2")))
    ws.conditional_formatting.add(f"J2:J{ws.max_row}", FormulaRule(formula=[f'LEN($J2)>0'], stopIfTrue=False,
                                                                   fill=PatternFill("solid", fgColor="ECFDF5")))

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"employees_{ts}.xlsx")


@app.route('/landing')
def landing():
    return render_template("landing.html")


# --- API & DEBUG ROUTES ---
@app.post("/api/bot/chats/recheck")
def api_bot_chats_recheck():
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    token = get_config_value("BOT_TOKEN")
    if not token:
        return jsonify({"ok": False, "error": "BOT_TOKEN is not set in DB"}), 400

    async def _recheck():
        bot = Bot(token=token, default=DefaultBotProperties(parse_mode="HTML"))
        try:
            me = await bot.get_me()
            updated, errors = 0, []
            with get_session() as s:
                rows = s.execute(select(GroupChat)).scalars().all()
                for r in rows:
                    try:
                        sid = int(str(r.chat_id).strip())
                        try:
                            m = await bot.get_chat_member(sid, me.id)
                        except TelegramBadRequest as e1:
                            if "chat not found" in str(e1).lower() and r.username:
                                try:
                                    ch2 = await bot.get_chat("@" + r.username.lstrip("@"))
                                    sid = ch2.id
                                    r.chat_id = sid
                                    m = await bot.get_chat_member(sid, me.id)
                                except Exception as e2:
                                    errors.append({"chat_id": r.chat_id, "error": f"fallback failed: {e2}"})
                                    continue
                            else:
                                errors.append({"chat_id": r.chat_id, "error": str(e1)})
                                continue
                        status_value = getattr(getattr(m, "status", None), "value", getattr(m, "status", None))
                        is_admin = status_value in ("creator", "administrator")
                        ch = await bot.get_chat(sid)
                        r.is_admin = bool(is_admin)
                        r.title = getattr(ch, "title", None) or getattr(ch, "full_name", None) or str(sid)
                        r.username = getattr(ch, "username", "")
                        r.type = getattr(getattr(ch, "type", None), "value", getattr(ch, "type", None)) or ""
                        r.updated_at = datetime.utcnow()
                        s.add(r)
                        updated += 1
                    except (TelegramBadRequest, TelegramForbiddenError) as e:
                        errors.append({"chat_id": r.chat_id, "error": str(e)})
                s.commit()
            return {"updated": updated, "errors": errors, "bot": {"id": me.id, "username": me.username}}
        finally:
            await bot.session.close()

    result = asyncio.run(_recheck())
    return jsonify({"ok": True, **result})


@app.route("/api/bot/chats", methods=["GET"])
def api_bot_chats_list():
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    with get_session() as db:
        rows = db.query(GroupChat).order_by(GroupChat.updated_at.desc().nullslast()).all()
        data = [{
            "chat_id": r.chat_id, "title": r.title or "", "username": r.username or "",
            "type": r.type or "", "is_admin": bool(r.is_admin),
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        } for r in rows]
        return jsonify({"ok": True, "data": data})


@app.route('/debug/me')
@login_required
def debug_me():
    async def _me():
        token = get_config_value("BOT_TOKEN")
        if not token:
            return {"error": "BOT_TOKEN not set in DB"}
        bot = Bot(token=token)
        try:
            me = await bot.get_me()
            return {"id": me.id, "username": me.username, "name": me.first_name}
        finally:
            await bot.session.close()

    try:
        info = asyncio.run(_me())
    except RuntimeError:
        info = {"note": "event loop busy; re-run or check logs"}
    return jsonify(info)


@app.route('/debug/diag_chat')
@login_required
def debug_diag_chat():
    raw = str(request.args.get('chat_id') or "").strip()
    if not raw:
        return jsonify(ok=False, error="no chat_id provided")

    async def _diag(raw_id: str):
        token = get_config_value("BOT_TOKEN")
        if not token:
            return {"ok": False, "error": "BOT_TOKEN not set in DB"}
        bot = Bot(token=token)
        try:
            me = await bot.get_me()
            out = {"me": {"id": me.id, "username": me.username}, "candidates": _chat_candidates(raw_id), "results": []}
            for cid in out["candidates"]:
                item = {"cid": cid}
                try:
                    chat = await bot.get_chat(cid)
                    item.update({"resolved_id": chat.id, "title": getattr(chat, "title", None),
                                 "type": getattr(chat, "type", None)})
                    try:
                        mem = await bot.get_chat_member(chat.id, me.id)
                        item["bot_status"] = str(mem.status)
                    except Exception as e:
                        item["bot_status_error"] = str(e)
                except Exception as e:
                    item["error"] = str(e)
                out["results"].append(item)
            return out
        finally:
            await bot.session.close()

    try:
        res = asyncio.run(_diag(raw))
    except RuntimeError:
        return jsonify(ok=False, error="event loop busy, retry once"), 503
    return jsonify(ok=True, **res)


@app.route("/circle/upload", methods=["POST"])
@login_required
def upload_circle_video():
    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"success": False, "message": "Файл не передан", "category": "danger"}), 400

    # Читаем в память
    file_bytes = file.read()
    mime_type = file.mimetype
    original_name = secure_filename(file.filename)

    uploaded_by = session.get("admin_username") or "admin"

    # пишем в БД
    with get_session() as db:
        circle = CircleVideo(
            file_data=file_bytes,
            file_mime=mime_type,
            original_filename=original_name,
            uploaded_by=uploaded_by,
        )
        db.add(circle)
        db.commit()  # commit, чтобы получить circle.id

    return jsonify({
        "success": True,
        "message": "Видео успешно загружено в БД.",
        "category": "success",
        "item": {
            "id": circle.id,
            # "stored_filename": circle.stored_filename, <-- Убрали
            "original_filename": circle.original_filename,
            # Обновляем URL, чтобы он вёл на новый маршрут с ID
            "url": url_for("serve_circle_video", circle_id=circle.id)
        }
    })


# --- НОВЫЕ МАРШРУТЫ ДЛЯ ОТДАЧИ ФАЙЛОВ ИЗ БД ---

def _serve_file_from_db(model_class, file_id: int):
    """Хелпер для отдачи файла из БД по ID."""
    with get_session() as db:
        item = db.get(model_class, file_id)

        if not item:
            return "File not found by ID", 404

        file_data, mime, name = None, None, None

        # Проверяем атрибуты для разных моделей
        if hasattr(item, 'file_data') and item.file_data:
            file_data = item.file_data
            mime = getattr(item, 'file_mime', 'application/octet-stream')
            name = getattr(item, 'file_name', 'download')
        elif hasattr(item, 'image_data') and item.image_data:  # для Topic
            file_data = item.image_data
            mime = getattr(item, 'image_mime', 'image/jpeg')
            name = getattr(item, 'image_name', 'image.jpg')
        elif hasattr(item, 'original_filename'):  # для CircleVideo
            name = item.original_filename

        if not file_data:
            return "File data not found in record", 404

        return send_file(
            io.BytesIO(file_data),
            mimetype=mime,
            download_name=name,
            as_attachment=False  # Показываем в браузере, если возможно
        )


@app.route("/circle/files/<int:circle_id>")
@login_required
def serve_circle_video(circle_id):
    # Старый маршрут /circle/files/<path:filename> больше не нужен
    return _serve_file_from_db(CircleVideo, circle_id)


@app.route("/files/topic_image/<int:topic_id>")
@login_required
def serve_topic_image(topic_id):
    return _serve_file_from_db(Topic, topic_id)


@app.route("/files/onboarding_step/<int:step_id>")
@login_required
def serve_onboarding_step_file(step_id):
    return _serve_file_from_db(OnboardingStep, step_id)


@app.route("/files/role_guide/<int:guide_id>")
@login_required
def serve_role_guide_file(guide_id):
    return _serve_file_from_db(RoleGuide, guide_id)


@app.route("/files/role_onboarding/<int:onboarding_id>")
@login_required
def serve_role_onboarding_file(onboarding_id):
    return _serve_file_from_db(RoleOnboarding, onboarding_id)


# --- ЗАПУСК ПРИЛОЖЕНИЯ ---
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)